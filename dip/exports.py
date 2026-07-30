"""Executive deliverables generated from the same plan the dashboard shows.

- :func:`build_pdf`   -> multi-page executive review (matplotlib ``PdfPages``)
- :func:`build_excel` -> multi-sheet workbook (openpyxl)

Both take an in-memory buffer so the Flask endpoints can stream them and the
CLI script can write them to ``deliverables/``. Both accept the dataset to
report on (``ds``): the Flask endpoints pass the *currently served* dataset, so
after an Excel import the deliverables reflect the imported data — the same
numbers as the screen. Both also carry the plan explanation (a "Why this plan?"
PDF page / an Explanation workbook sheet), rendered from the single structure
:func:`dip.explain.explain_plan` builds.
"""

from __future__ import annotations

import io

import matplotlib

matplotlib.use("Agg")  # headless: safe in CI / containers
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

from .analytics import (
    abc_xyz,
    kpi_drilldown,
    kpis,
    margin_bridge,
    revenue_breakdown,
    rfm_segments,
)
from .crosssell import mine_crosssell
from .data import Dataset, build_dataset
from .explain import explain_plan
from .forecast import forecast_revenue
from .optimize import optimize_assortment, optimize_prices, optimize_routes
from .prescribe import build_plan
from .scenario import DEFAULT_MAX_CHANGE, KPI_KEYS, compare_scenarios

INK = "#1a2233"
BLUE = "#2f6bff"
GREEN = "#1d9e6f"
PINK = "#ea4b71"
GREY = "#8a93a6"


def _eur(x: float) -> str:
    return f"EUR {x:,.0f}"


def _scenario_line(plan: dict) -> str:
    """One line naming the scenario a deliverable was computed for."""
    budget_pct = plan["budget"] / plan["full_capital"] if plan.get("full_capital") else 0.0
    return (
        f"Scenario: working-capital budget {_eur(plan['budget'])} of {_eur(plan['full_capital'])} "
        f"({budget_pct:.0%}), price guardrail +/-{plan['max_change']:.0%}"
    )


def _source_line(source: dict | None) -> str:
    """The deliverable's data-source disclosure (honest in both modes)."""
    if source is None or source.get("synthetic", True):
        return "Synthetic data - generated deterministically. Author: Dimitres Kisimov, 2026."
    return (
        f"Your data: {source.get('filename')} ({source.get('n_skus')} SKUs, imported; "
        "customers & routing remain synthetic). Author: Dimitres Kisimov, 2026."
    )


def _plan_and_explanation(
    ds: Dataset,
    budget: float | None,
    max_change: float,
    plan: dict | None,
    explanation: dict | None,
    source: dict | None,
) -> tuple[dict, dict]:
    """Resolve the (plan, explanation) pair for a deliverable.

    Callers that hold both (the Flask endpoints, the CLI) pass them through
    untouched. Otherwise one routing + pricing solve is shared by the plan and
    its explanation; every engine is deterministic, so a recompute can never
    disagree with a passed-in plan.
    """
    if plan is not None and explanation is not None:
        return plan, explanation
    routes = optimize_routes(ds)
    prices = optimize_prices(ds, max_change=max_change)
    if plan is None:
        plan = build_plan(ds, budget=budget, max_change=max_change, routes=routes, prices=prices)
    if explanation is None:
        explanation = explain_plan(
            ds, max_change=max_change, plan=plan, routes=routes, prices=prices, source=source
        )
    return plan, explanation


def build_pdf(
    budget: float | None = None,
    max_change: float = 0.15,
    plan: dict | None = None,
    ds: Dataset | None = None,
    source: dict | None = None,
    explanation: dict | None = None,
) -> bytes:
    """Render the executive-review PDF for the given scenario and return its bytes.

    ``budget`` / ``max_change`` mirror the dashboard controls so the exported
    deck matches what was on screen when the user clicked Export. Callers that
    already hold the plan for this exact scenario (the Flask endpoints pass the
    served one, the CLI builds one and shares it) pass it as ``plan`` so the
    deliverable quotes *the same object* the dashboard shows — never a re-solve.
    ``ds`` is the dataset to report on (the Flask endpoints pass the currently
    served one, so imported data flows into the deck); ``explanation`` the
    matching :func:`dip.explain.explain_plan` structure for the "Why" page.
    """
    ds = ds if ds is not None else build_dataset()
    k = kpis(ds)
    fc = forecast_revenue(ds)
    plan, explanation = _plan_and_explanation(ds, budget, max_change, plan, explanation, source)
    mb = margin_bridge(ds)
    rb = revenue_breakdown(ds)
    az = abc_xyz(ds)

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        # ---- Page 1: cover + KPIs ----------------------------------------
        fig = plt.figure(figsize=(11.69, 8.27))  # A4 landscape
        fig.patch.set_facecolor("white")
        fig.text(0.06, 0.90, "Distributor Intelligence Platform", fontsize=26, weight="bold", color=INK)
        fig.text(0.06, 0.855, "Executive decision review", fontsize=14, color=GREY)
        fig.text(0.06, 0.80, f"Expected annual uplift: {_eur(plan['expected_uplift_eur'])} "
                             f"({plan['expected_uplift_pct']:.1%} of annual gross margin)",
                 fontsize=15, color=GREEN, weight="bold")
        fig.text(0.06, 0.765, _scenario_line(plan), fontsize=10.5, color=GREY)

        tiles = [
            ("Revenue (24 mo)", _eur(k["revenue"])),
            ("Gross margin", f"{k['gross_margin_pct']:.1%}"),
            ("YoY growth", f"{k['yoy']:+.1%}"),
            ("Forecast MASE", f"{fc['mase']}"),
            ("OTIF service", f"{k['otif']:.1%}"),
            ("SKUs / customers", f"{k['n_skus']} / {k['n_customers']}"),
        ]
        for i, (lab, val) in enumerate(tiles):
            x = 0.06 + (i % 3) * 0.31
            y = 0.60 - (i // 3) * 0.22
            fig.text(x, y, val, fontsize=20, weight="bold", color=BLUE)
            fig.text(x, y - 0.05, lab, fontsize=11, color=GREY)
        fig.text(0.06, 0.06, _source_line(source), fontsize=8, color=GREY)
        pdf.savefig(fig)
        plt.close(fig)

        # ---- Page 2: revenue history + forecast --------------------------
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        hist_x = list(range(len(fc["history"])))
        fut_x = list(range(len(fc["history"]), len(fc["history"]) + len(fc["forecast"])))
        ax.plot(hist_x, fc["history"], color=BLUE, lw=2.2, label="Actual revenue")
        ax.plot(fut_x, fc["forecast"], color=GREEN, lw=2.2, ls="--", label="Forecast")
        ax.fill_between(fut_x, fc["lower"], fc["upper"], color=GREEN, alpha=0.15, label="~80% band")
        ax.set_title(f"Monthly revenue and forecast  |  {fc['method']}  |  MASE {fc['mase']}",
                     fontsize=14, color=INK, weight="bold")
        ax.set_ylabel("EUR / month")
        ax.legend(frameon=False)
        ax.spines[["top", "right"]].set_visible(False)
        pdf.savefig(fig)
        plt.close(fig)

        # ---- Page 3: margin bridge + revenue by region -------------------
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11.69, 8.27))
        # waterfall
        labels = ["Baseline"] + [s["label"] for s in mb["steps"]] + ["Current"]
        vals = [mb["baseline"]] + [s["value"] for s in mb["steps"]] + [mb["current"]]
        running = mb["baseline"]
        for i, (lab, v) in enumerate(zip(labels, vals)):
            if i == 0 or i == len(labels) - 1:
                ax1.bar(i, v, color=INK)
            else:
                color = GREEN if v >= 0 else PINK
                ax1.bar(i, v, bottom=running, color=color)
                running += v
        ax1.set_xticks(range(len(labels)))
        ax1.set_xticklabels(labels, rotation=20, ha="right", fontsize=9)
        ax1.set_title("Gross-margin bridge (H1 -> H2)", fontsize=12, color=INK, weight="bold")
        ax1.spines[["top", "right"]].set_visible(False)
        # region bars
        regs = rb["region"]
        ax2.barh([r["label"] for r in regs], [r["revenue"] for r in regs], color=BLUE)
        ax2.invert_yaxis()
        ax2.set_title("Revenue by region", fontsize=12, color=INK, weight="bold")
        ax2.spines[["top", "right"]].set_visible(False)
        pdf.savefig(fig)
        plt.close(fig)

        # ---- Page 4: recommended actions ---------------------------------
        fig = plt.figure(figsize=(11.69, 8.27))
        fig.text(0.06, 0.92, "Recommended actions", fontsize=20, weight="bold", color=INK)
        y = 0.82
        for c in plan["cards"]:
            fig.text(0.06, y, f"[{c['lever']}]  {c['title']}", fontsize=13, weight="bold", color=BLUE)
            fig.text(0.06, y - 0.035, c["detail"], fontsize=10, color=INK, wrap=True)
            fig.text(0.82, y, _eur(c["impact_eur"]), fontsize=12, weight="bold", color=GREEN)
            y -= 0.13
        _ = az  # abc-xyz already summarised in workbook; keep import meaningful
        pdf.savefig(fig)
        plt.close(fig)

        # ---- Page 5: why this plan? (rendered from dip.explain's structure) --
        ex = explanation
        fig = plt.figure(figsize=(11.69, 8.27))
        fig.text(0.06, 0.93, "Why this plan?", fontsize=20, weight="bold", color=INK)
        fig.text(
            0.06, 0.885,
            f"Expected uplift {_eur(ex['headline']['expected_uplift_eur'])}/yr "
            f"({ex['headline']['expected_uplift_pct']:.1%}) vs {ex['headline']['baseline']}.",
            fontsize=11, color=INK,
        )
        y = 0.835
        fig.text(0.06, y, "Binding constraints", fontsize=12.5, weight="bold", color=BLUE)
        y -= 0.033
        for bc in ex["binding_constraints"]:
            tag = "BINDING" if bc["binding"] else "slack"
            fig.text(0.06, y, f"[{tag}] {bc['detail']}", fontsize=9, color=INK)
            y -= 0.028
        y -= 0.012
        fig.text(0.06, y, "What changes vs doing nothing", fontsize=12.5, weight="bold", color=BLUE)
        y -= 0.033
        for lv in ex["levers"]:
            fig.text(0.06, y, f"{lv['lever'].title()} — {_eur(lv['total_eur'])}/yr",
                     fontsize=10.5, weight="bold", color=INK)
            y -= 0.026
            for mv in lv["moves"]:
                who = f"{mv['sku_id']}: " if mv["sku_id"] else ""
                fig.text(0.08, y, f"- {who}{mv['action']}  ({_eur(mv['contribution_eur'])})",
                         fontsize=8.5, color=INK)
                y -= 0.023
            fig.text(0.08, y, lv["note"], fontsize=8, color=GREY)
            y -= 0.029
        sens = ex["sensitivity"]
        fig.text(0.06, y, "Sensitivity (budget -10% / +10%)", fontsize=12.5, weight="bold", color=BLUE)
        y -= 0.030
        fig.text(
            0.06, y,
            f"At {_eur(sens['budget_minus_10pct']['budget'])}: uplift "
            f"{_eur(sens['budget_minus_10pct']['expected_uplift_eur'])} "
            f"({sens['budget_minus_10pct']['delta_eur']:+,.0f}); at "
            f"{_eur(sens['budget_plus_10pct']['budget'])}: "
            f"{_eur(sens['budget_plus_10pct']['expected_uplift_eur'])} "
            f"({sens['budget_plus_10pct']['delta_eur']:+,.0f}). {sens['note']}.",
            fontsize=9, color=INK,
        )
        y -= 0.036
        fig.text(0.06, y, "Caveats", fontsize=12.5, weight="bold", color=BLUE)
        y -= 0.030
        for cv in ex["caveats"]:
            fig.text(0.06, y, f"- {cv}", fontsize=8, color=GREY)
            y -= 0.023
        pdf.savefig(fig)
        plt.close(fig)

    return buf.getvalue()


def build_excel(
    budget: float | None = None,
    max_change: float = 0.15,
    plan: dict | None = None,
    ds: Dataset | None = None,
    source: dict | None = None,
    explanation: dict | None = None,
    scenario_compare: dict | None = None,
) -> bytes:
    """Render the multi-sheet workbook for the given scenario and return its bytes.

    ``plan``, when supplied, must be the plan built for this exact scenario —
    it is reused verbatim (see :func:`build_pdf`) instead of being re-solved.
    ``ds`` / ``source`` / ``explanation`` mirror :func:`build_pdf`: the workbook
    reports on the currently served dataset (imported or synthetic) and carries
    an Explanation sheet rendered from the same structure as ``/api/explain``.
    ``scenario_compare`` is the :func:`dip.scenario.compare_scenarios` A/B
    structure rendered on the Scenarios sheet; when omitted it is computed here
    as the default baseline vs the on-screen scenario, so the sheet's numbers
    match what ``POST /api/scenario/compare`` returns for those two scenarios
    (screen == export).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    ds = ds if ds is not None else build_dataset()
    k = kpis(ds)
    fc = forecast_revenue(ds)
    plan, explanation = _plan_and_explanation(ds, budget, max_change, plan, explanation, source)
    if scenario_compare is None:
        scenario_compare = compare_scenarios(
            ds,
            scenario_a={
                "name": "Baseline (default budget & guardrail)",
                "budget": None,
                "max_change": DEFAULT_MAX_CHANGE,
            },
            scenario_b={"name": "On-screen scenario", "budget": budget, "max_change": max_change},
            source=source,
        )
    rb = revenue_breakdown(ds)
    az = abc_xyz(ds)
    assort = optimize_assortment(ds, budget=budget)
    prices = optimize_prices(ds, max_change=max_change)
    rfm = rfm_segments(ds)

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="2F6BFF")
    head_font = Font(bold=True, color="FFFFFF")

    def _header(ws, row: int, cols: list[str]):
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=j, value=c)
            cell.fill = head_fill
            cell.font = head_font

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Distributor Intelligence Platform - Executive Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = _scenario_line(plan)
    ws["A2"].font = Font(italic=True, size=10, color="6B7488")
    rows = [
        ("Revenue (24 mo)", k["revenue"]),
        ("Gross margin", k["gross_margin"]),
        ("Gross margin %", k["gross_margin_pct"]),
        ("YoY growth", k["yoy"]),
        ("OTIF service (modelled)", k["otif"]),
        ("Forecast MASE", fc["mase"]),
        ("Scenario: working-capital budget (EUR)", plan["budget"]),
        ("Scenario: price guardrail (+/-)", plan["max_change"]),
        ("Expected annual uplift (EUR)", plan["expected_uplift_eur"]),
        ("Uplift % of annual gross margin", plan["expected_uplift_pct"]),
        (
            "Data source",
            "seeded synthetic dataset"
            if source is None or source.get("synthetic", True)
            else f"imported: {source.get('filename')} ({source.get('n_skus')} SKUs)",
        ),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for i, (m, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20

    # Forecast
    ws = wb.create_sheet("Forecast")
    _header(ws, 1, ["Month", "Type", "Revenue", "Lower", "Upper"])
    r = 2
    for m, v in zip(fc["history_months"], fc["history"]):
        ws.cell(row=r, column=1, value=m); ws.cell(row=r, column=2, value="actual")
        ws.cell(row=r, column=3, value=v); r += 1
    for m, v, lo, hi in zip(fc["forecast_months"], fc["forecast"], fc["lower"], fc["upper"]):
        ws.cell(row=r, column=1, value=m); ws.cell(row=r, column=2, value="forecast")
        ws.cell(row=r, column=3, value=v); ws.cell(row=r, column=4, value=lo)
        ws.cell(row=r, column=5, value=hi); r += 1

    # Revenue breakdown
    ws = wb.create_sheet("Revenue")
    _header(ws, 1, ["Dimension", "Label", "Revenue"])
    r = 2
    for dim in ("region", "category", "channel"):
        for row in rb[dim]:
            ws.cell(row=r, column=1, value=dim); ws.cell(row=r, column=2, value=row["label"])
            ws.cell(row=r, column=3, value=row["revenue"]); r += 1

    # ABC-XYZ
    ws = wb.create_sheet("ABC-XYZ")
    _header(ws, 1, ["Cell", "SKU count", "Revenue"])
    for r, (cell, v) in enumerate(sorted(az["grid"].items()), start=2):
        ws.cell(row=r, column=1, value=cell); ws.cell(row=r, column=2, value=v["count"])
        ws.cell(row=r, column=3, value=v["revenue"])

    # Assortment
    ws = wb.create_sheet("Assortment")
    _header(ws, 1, ["Plan", "SKUs carried", "Margin", "Capital used"])
    ws.append(["MILP (optimal)", assort["milp"]["count"], assort["milp"]["margin"], assort["milp"]["capital_used"]])
    ws.append(["Greedy baseline", assort["greedy"]["count"], assort["greedy"]["margin"], assort["greedy"]["capital_used"]])
    ws.append(["Uplift", "", assort["uplift_vs_greedy"], ""])

    # Pricing (top elasticity-guided moves by profit delta)
    ws = wb.create_sheet("Pricing")
    _header(ws, 1, ["SKU", "Category", "Price old", "Price new", "Change %", "Elasticity", "Profit delta (EUR)"])
    for r, rec in enumerate(prices["recommendations"][:40], start=2):
        ws.cell(row=r, column=1, value=rec["sku_id"]); ws.cell(row=r, column=2, value=rec["category"])
        ws.cell(row=r, column=3, value=rec["price_old"]); ws.cell(row=r, column=4, value=rec["price_new"])
        ws.cell(row=r, column=5, value=rec["change_pct"]); ws.cell(row=r, column=6, value=rec["elasticity"])
        ws.cell(row=r, column=7, value=rec["profit_delta"])
    ws.column_dimensions["B"].width = 16

    # RFM segments
    ws = wb.create_sheet("RFM")
    _header(ws, 1, ["Segment", "Customers", "Monetary (EUR)"])
    for r, seg in enumerate(rfm["segments"], start=2):
        ws.cell(row=r, column=1, value=seg["segment"]); ws.cell(row=r, column=2, value=seg["count"])
        ws.cell(row=r, column=3, value=seg["monetary"])
    ws.column_dimensions["A"].width = 24

    # KPI drill-downs — the same rows the dashboard's headline tiles open;
    # both are computed by dip.analytics.kpi_drilldown, so they cannot differ.
    kd = kpi_drilldown(ds)
    ws = wb.create_sheet("Drill-downs")
    ws["A1"] = "Revenue by segment (region x channel)"
    ws["A1"].font = Font(bold=True, size=12)
    _header(ws, 2, ["Region", "Channel", "Revenue (EUR)", "Share"])
    r = 3
    for row in kd["revenue_by_segment"]["rows"]:
        ws.cell(row=r, column=1, value=row["region"]); ws.cell(row=r, column=2, value=row["channel"])
        ws.cell(row=r, column=3, value=row["revenue"]); ws.cell(row=r, column=4, value=row["share"])
        r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=r, column=3, value=kd["revenue_by_segment"]["total_revenue"]).font = Font(bold=True)
    r += 2
    ws.cell(row=r, column=1, value="Gross margin by category").font = Font(bold=True, size=12)
    r += 1
    _header(ws, r, ["Category", "Revenue (EUR)", "COGS (EUR)", "Gross margin (EUR)", "Margin %", "Share of margin"])
    for row in kd["margin_by_category"]["rows"]:
        r += 1
        ws.cell(row=r, column=1, value=row["category"]); ws.cell(row=r, column=2, value=row["revenue"])
        ws.cell(row=r, column=3, value=row["cogs"]); ws.cell(row=r, column=4, value=row["gross_margin"])
        ws.cell(row=r, column=5, value=row["margin_pct"]); ws.cell(row=r, column=6, value=row["share_of_margin"])
    r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=r, column=2, value=kd["margin_by_category"]["total_revenue"]).font = Font(bold=True)
    ws.cell(row=r, column=3, value=kd["margin_by_category"]["total_cogs"]).font = Font(bold=True)
    ws.cell(row=r, column=4, value=kd["margin_by_category"]["total_gross_margin"]).font = Font(bold=True)
    r += 2
    ws.cell(row=r, column=1, value=kd["note"]).font = Font(italic=True, size=9, color="6B7488")
    for col, width in (("A", 18), ("B", 14), ("C", 14), ("D", 18), ("E", 10), ("F", 14)):
        ws.column_dimensions[col].width = width

    # Cross-sell rules — deterministic mining, so this sheet quotes the exact
    # numbers /api/crosssell (and the dashboard card) serve for this dataset.
    cs = mine_crosssell(ds)
    ws = wb.create_sheet("Cross-sell")
    ws["A1"] = "Cross-sell association rules (Apriori)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = cs["note"]
    ws["A2"].font = Font(italic=True, size=9, color="6B7488")
    if not cs["available"]:
        ws["A4"] = "No basket data in this dataset."
    else:
        ws["A3"] = (
            f"{cs['n_rules']} rules from {cs['n_baskets']} baskets "
            f"(min support {cs['params']['min_support']}, min confidence "
            f"{cs['params']['min_confidence']}, min lift {cs['params']['min_lift']}); top 40 by lift below. "
            "Rules on few baskets are flagged thin."
        )
        ws["A3"].font = Font(size=9, color="6B7488")
        _header(ws, 5, ["If they buy", "Name", "Also sells", "Name", "Support", "Confidence", "Lift", "Baskets", "Thin support?"])
        for r, rule in enumerate(cs["rules"][:40], start=6):
            ws.cell(row=r, column=1, value=rule["antecedent"]); ws.cell(row=r, column=2, value=rule["antecedent_name"])
            ws.cell(row=r, column=3, value=rule["consequent"]); ws.cell(row=r, column=4, value=rule["consequent_name"])
            ws.cell(row=r, column=5, value=rule["support"]); ws.cell(row=r, column=6, value=rule["confidence"])
            ws.cell(row=r, column=7, value=rule["lift"]); ws.cell(row=r, column=8, value=rule["support_count"])
            ws.cell(row=r, column=9, value="thin" if rule["thin_support"] else "")
    for col, width in (("A", 12), ("B", 16), ("C", 12), ("D", 16), ("E", 10), ("F", 11), ("G", 8), ("H", 9), ("I", 12)):
        ws.column_dimensions[col].width = width

    # Actions
    ws = wb.create_sheet("Actions")
    _header(ws, 1, ["Lever", "Action", "Impact (EUR)", "Confidence"])
    for r, c in enumerate(plan["cards"], start=2):
        ws.cell(row=r, column=1, value=c["lever"]); ws.cell(row=r, column=2, value=c["title"])
        ws.cell(row=r, column=3, value=c["impact_eur"]); ws.cell(row=r, column=4, value=c["confidence"])
    ws.column_dimensions["B"].width = 44

    # Explanation — the same structure /api/explain and the PDF "Why" page use
    ex = explanation
    ws = wb.create_sheet("Explanation")
    ws["A1"] = "Why this plan?"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Expected uplift EUR {ex['headline']['expected_uplift_eur']:,.0f}/yr "
        f"({ex['headline']['expected_uplift_pct']:.1%}) vs {ex['headline']['baseline']}"
    )
    r = 4
    _header(ws, r, ["Constraint", "Binding?", "Detail"])
    for bc in ex["binding_constraints"]:
        r += 1
        ws.cell(row=r, column=1, value=bc["name"])
        ws.cell(row=r, column=2, value="BINDING" if bc["binding"] else "slack")
        ws.cell(row=r, column=3, value=bc["detail"])
    r += 2
    _header(ws, r, ["Lever", "Move", "Contribution (EUR)"])
    for lv in ex["levers"]:
        r += 1
        ws.cell(row=r, column=1, value=lv["lever"])
        ws.cell(row=r, column=2, value=f"lever total ({lv['n_moves']} moves)")
        ws.cell(row=r, column=3, value=lv["total_eur"])
        for mv in lv["moves"]:
            r += 1
            ws.cell(row=r, column=2, value=(f"{mv['sku_id']}: " if mv["sku_id"] else "") + mv["action"])
            ws.cell(row=r, column=3, value=mv["contribution_eur"])
        if lv["remainder_eur"]:
            r += 1
            ws.cell(row=r, column=2, value="all remaining moves together")
            ws.cell(row=r, column=3, value=lv["remainder_eur"])
    r += 2
    _header(ws, r, ["Sensitivity", "Budget (EUR)", "Uplift (EUR)", "Delta (EUR)"])
    for key, label in (("budget_minus_10pct", "budget -10%"), ("budget_plus_10pct", "budget +10%")):
        r += 1
        s = ex["sensitivity"][key]
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=s["budget"])
        ws.cell(row=r, column=3, value=s["expected_uplift_eur"])
        ws.cell(row=r, column=4, value=s["delta_eur"])
    r += 1
    ws.cell(row=r, column=1, value=ex["sensitivity"]["note"])
    r += 2
    _header(ws, r, ["Caveats"])
    for cv in ex["caveats"]:
        r += 1
        ws.cell(row=r, column=1, value=cv)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    # Scenarios — the A/B compare from dip.scenario.compare_scenarios, the exact
    # numbers POST /api/scenario/compare returns for the same two scenarios.
    sc = scenario_compare
    sa, sb = sc["scenario_a"], sc["scenario_b"]
    ws = wb.create_sheet("Scenarios")
    ws["A1"] = "Scenario compare (A vs B)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = sc["provenance"]
    ws["A2"].font = Font(italic=True, size=9, color="6B7488")
    ws["A3"] = f"A: {sa['name']}   |   B: {sb['name']}"
    ws["A3"].font = Font(size=10, color="6B7488")
    scenario_labels = [
        ("expected_uplift_eur", "Expected annual uplift (EUR)"),
        ("expected_uplift_pct", "Uplift % of annual gross margin"),
        ("pricing_uplift_eur", "Pricing lever (EUR)"),
        ("assortment_uplift_eur", "Assortment lever vs greedy (EUR)"),
        ("routing_uplift_eur", "Routing lever (EUR)"),
        ("margin_captured_eur", "Assortment margin captured (EUR)"),
        ("capital_used_eur", "Working capital used (EUR)"),
        ("skus_carried", "SKUs carried"),
        ("budget_eur", "Working-capital budget (EUR)"),
    ]
    # the sheet renders the full compared KPI set, in the module's canonical order
    assert [key for key, _label in scenario_labels] == list(KPI_KEYS)
    _header(ws, 5, ["Metric", "Scenario A", "Scenario B", "Delta (B-A)", "Delta %"])
    r = 6
    for key, label in scenario_labels:
        d = sc["deltas"][key]
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=sa["kpis"][key])
        ws.cell(row=r, column=3, value=sb["kpis"][key])
        ws.cell(row=r, column=4, value=d["abs"])
        ws.cell(row=r, column=5, value=d["pct"])
        r += 1
    r += 1
    _header(ws, r, ["Scenario", "Budget (EUR)", "Budget % of capital", "Price guardrail"])
    for side in (sa, sb):
        r += 1
        ws.cell(row=r, column=1, value=side["name"])
        ws.cell(row=r, column=2, value=side["budget"])
        ws.cell(row=r, column=3, value=side["budget_pct_of_full"])
        ws.cell(row=r, column=4, value=side["max_change"])
    r += 2
    ws.cell(row=r, column=1, value=sc["note"]).font = Font(italic=True, size=9, color="6B7488")
    ws.column_dimensions["A"].width = 34
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
