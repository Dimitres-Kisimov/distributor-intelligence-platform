"""Excel round-trip: import a user's own SKU workbook, run the whole platform on it.

The template (served by ``GET /api/import/template``, built by
:func:`build_template`) is **wide format** — the friendlier of the two layouts
for spreadsheet users, because a SKU is one row you can eyeball:

``SKUs`` sheet, one row per SKU
    ``sku, name, category, unit_cost, unit_price`` followed by 24 month
    columns (``YYYY-MM``, the same 24-month horizon the platform models).
    Each month cell holds the units sold that month; blank cells mean 0.

An ``Instructions`` sheet in the template documents every rule below, so the
workbook is self-describing.

Honesty contract (mirrors the logistics-digital-twin importer): only products
and their monthly units are imported. The customer base and the delivery-routing
layer stay synthetic (the template does not cover them), elasticities are
category defaults, and every analysis discloses this via the ``source`` payload
and the plan explanation's caveats. Imported data lives in process memory only —
nothing is written to disk; restarting the app restores the synthetic dataset.

Validation is strict and cell-addressed (``SKUs!D5: ...``) so a failed upload
is fixable without guesswork. A failed import never touches the running state.
"""

from __future__ import annotations

import io
import math

from .data import CATEGORIES, N_MONTHS, Dataset, _month_labels, build_dataset

# ---- documented limits ------------------------------------------------------
MAX_XLSX_BYTES = 2_000_000  # 2 MB upload cap — thousands of SKU rows fit easily
MAX_SKUS = 2000  # keeps the MILP + pricing engines interactive
MAX_ERRORS = 40  # errors reported per upload before truncating

# Elasticity used when the imported category doesn't match a known one.
DEFAULT_ELASTICITY = -1.2
# Fixed, documented defaults for fields the template intentionally omits.
DEFAULT_LEAD_TIME_DAYS = 10
DEFAULT_SHELF_SPACE = 1.0
IMPORTED_LABEL = "Imported"

FIXED_COLUMNS = ("sku", "name", "category", "unit_cost", "unit_price")


class XlsxImportError(ValueError):
    """A user-facing problem with an uploaded workbook.

    ``errors`` carries the individual row/cell-addressed messages; ``str(exc)``
    is the one-line summary.
    """

    def __init__(self, summary: str, errors: list[str] | None = None):
        super().__init__(summary)
        self.errors = errors or [summary]


def month_columns() -> list[str]:
    """The 24 ``YYYY-MM`` month column labels the template uses (in order)."""
    return _month_labels(N_MONTHS)


def _col_letter(idx: int) -> str:
    """1-based column index -> Excel letter (1 -> A, 27 -> AA)."""
    out = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def build_template() -> bytes:
    """An empty import template: Instructions sheet + SKUs sheet with headers."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    months = month_columns()
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("Distributor Intelligence Platform — import template", True),
        ("", False),
        ("Fill the SKUs sheet, one row per product, then upload the file", False),
        ("via 'Import Excel' on the dashboard (POST /api/import).", False),
        ("", False),
        ("Required columns (do not rename or reorder):", True),
        ("  sku         unique product code (text)", False),
        ("  name        product name (blank = the sku code is used)", False),
        ("  category    free text (blank = 'Uncategorised')", False),
        ("  unit_cost   cost per unit, >= 0", False),
        ("  unit_price  selling price per unit, > 0", False),
        (f"  {months[0]} .. {months[-1]}   units sold per month (24 columns).", False),
        ("              Numbers >= 0; blank cells count as 0.", False),
        ("", False),
        ("Rules:", True),
        (f"  - at most {MAX_SKUS} SKU rows; file at most {MAX_XLSX_BYTES // 1_000_000} MB", False),
        ("  - duplicate sku codes are rejected (case-insensitive)", False),
        ("  - keep the month columns exactly as generated here", False),
        ("  - extra columns are rejected so header typos are caught, not ignored", False),
        ("", False),
        ("What the platform assumes about imported data (disclosed honestly):", True),
        ("  - price elasticity: category default if the category matches one of", False),
        (f"    {', '.join(CATEGORIES)}; otherwise {DEFAULT_ELASTICITY}", False),
        (f"  - lead time {DEFAULT_LEAD_TIME_DAYS} days and shelf space {DEFAULT_SHELF_SPACE} per SKU", False),
        ("    (the template does not collect them)", False),
        ("  - customers and delivery routing stay the synthetic demo layer:", False),
        ("    the template covers products only", False),
        ("", False),
        ("Your data is held in the app's memory only — never written to disk.", False),
        ("Restarting the app (or POST /api/reset) restores the synthetic dataset.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 78

    ws = wb.create_sheet("SKUs")
    head_fill = PatternFill("solid", fgColor="2F6BFF")
    head_font = Font(bold=True, color="FFFFFF")
    for j, col in enumerate(list(FIXED_COLUMNS) + months, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = head_fill
        cell.font = head_font
    for j, width in enumerate((14, 24, 16, 11, 11), start=1):
        ws.column_dimensions[_col_letter(j)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Parsing + validation
# ---------------------------------------------------------------------------


def _number(value, allow_blank: bool = False) -> float | None:
    """Parse a cell into a finite float; ``None`` signals 'not a number'."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return 0.0 if allow_blank else None
    if isinstance(value, bool):  # Excel TRUE/FALSE must not sneak in as 1/0
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return v if math.isfinite(v) else None
    try:
        v = float(str(value).strip())
    except ValueError:
        return None
    return v if math.isfinite(v) else None


def parse_workbook(data: bytes, filename: str = "upload.xlsx") -> list[dict]:
    """Parse + validate an uploaded workbook into SKU row dicts.

    Raises :class:`XlsxImportError` (with cell-addressed ``errors``) on any
    problem; returns clean rows otherwise. Each row dict carries ``sku``,
    ``name``, ``category``, ``unit_cost``, ``unit_price`` and ``monthly_units``
    (a list of 24 floats aligned with :func:`month_columns`).
    """
    from openpyxl import load_workbook

    if len(data) > MAX_XLSX_BYTES:
        raise XlsxImportError(
            f"{filename}: file exceeds the {MAX_XLSX_BYTES // 1_000_000} MB limit"
        )
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several zip/xml error types here
        raise XlsxImportError(f"{filename}: not a readable .xlsx workbook ({exc})") from exc

    try:
        if "SKUs" not in wb.sheetnames:
            raise XlsxImportError(
                f"{filename}: missing the 'SKUs' sheet (found: {', '.join(wb.sheetnames) or 'none'}). "
                "Start from GET /api/import/template."
            )
        ws = wb["SKUs"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise XlsxImportError(f"{filename}: the SKUs sheet is empty")

        expected = list(FIXED_COLUMNS) + month_columns()
        got = [str(c).strip() if c is not None else "" for c in header]
        errors: list[str] = []
        for j, want in enumerate(expected):
            have = got[j] if j < len(got) else ""
            if have != want:
                errors.append(
                    f"SKUs!{_col_letter(j + 1)}1: expected column '{want}', got '{have or '(blank)'}'"
                )
        for j in range(len(expected), len(got)):
            if got[j]:
                errors.append(f"SKUs!{_col_letter(j + 1)}1: unexpected extra column '{got[j]}'")
        if errors:
            raise XlsxImportError(
                f"{filename}: header does not match the template", errors[:MAX_ERRORS]
            )

        n_months = len(month_columns())
        out: list[dict] = []
        seen: dict[str, int] = {}
        for r, raw in enumerate(rows_iter, start=2):
            raw = list(raw) + [None] * (len(expected) - len(raw))
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
                continue  # fully blank row: ignore
            sku = str(raw[0]).strip() if raw[0] is not None else ""
            if not sku:
                errors.append(f"SKUs!A{r}: sku is blank")
                continue
            if sku.upper() in seen:
                errors.append(
                    f"SKUs!A{r}: duplicate sku '{sku}' (first used in row {seen[sku.upper()]})"
                )
                continue
            seen[sku.upper()] = r

            name = str(raw[1]).strip() if raw[1] is not None else ""
            category = str(raw[2]).strip() if raw[2] is not None else ""
            cost = _number(raw[3])
            price = _number(raw[4])
            row_bad = False
            if cost is None:
                errors.append(f"SKUs!D{r}: unit_cost must be a number, got {raw[3]!r}")
                row_bad = True
            elif cost < 0:
                errors.append(f"SKUs!D{r}: unit_cost must be >= 0, got {cost}")
                row_bad = True
            if price is None:
                errors.append(f"SKUs!E{r}: unit_price must be a number, got {raw[4]!r}")
                row_bad = True
            elif price <= 0:
                # zero/negative price breaks the elasticity maths downstream —
                # rejected here rather than dividing by zero later.
                errors.append(f"SKUs!E{r}: unit_price must be > 0, got {price}")
                row_bad = True

            units: list[float] = []
            for j in range(n_months):
                col = len(FIXED_COLUMNS) + j
                u = _number(raw[col], allow_blank=True)
                addr = f"SKUs!{_col_letter(col + 1)}{r}"
                if u is None:
                    errors.append(f"{addr}: units must be a number, got {raw[col]!r}")
                    row_bad = True
                elif u < 0:
                    errors.append(f"{addr}: units must be >= 0, got {u}")
                    row_bad = True
                else:
                    units.append(u)
            if row_bad or len(errors) >= MAX_ERRORS:
                if len(errors) >= MAX_ERRORS:
                    break
                continue
            out.append(
                {
                    "sku": sku,
                    "name": name or sku,
                    "category": category or "Uncategorised",
                    "unit_cost": round(float(cost), 4),
                    "unit_price": round(float(price), 4),
                    "monthly_units": units,
                }
            )

        if errors:
            raise XlsxImportError(
                f"{filename}: {len(errors)} problem(s) found — nothing was imported",
                errors[:MAX_ERRORS],
            )
        if not out:
            raise XlsxImportError(f"{filename}: the SKUs sheet has no data rows")
        if len(out) > MAX_SKUS:
            raise XlsxImportError(
                f"{filename}: at most {MAX_SKUS} SKU rows are supported (got {len(out)})"
            )
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Dataset assembly
# ---------------------------------------------------------------------------

_ELASTICITY_BY_CATEGORY = {name.lower(): vals[2] for name, vals in CATEGORIES.items()}


def _elasticity_for(category: str) -> float:
    return float(_ELASTICITY_BY_CATEGORY.get(category.lower(), DEFAULT_ELASTICITY))


def dataset_from_rows(rows: list[dict]) -> tuple[Dataset, list[str]]:
    """Build a full :class:`~dip.data.Dataset` around imported SKU rows.

    Monthly facts come straight from the imported units x prices. The customer
    base, depot and route stops are reused from the seeded synthetic build (the
    template covers products only) — this is disclosed in the returned
    ``assumptions`` and in every plan explanation.
    """
    import numpy as np

    months = month_columns()
    synth = build_dataset()  # cached; supplies the customer/routing layer

    skus: list[dict] = []
    rows_month, rows_sku, rows_region = [], [], []
    rows_channel, rows_category = [], []
    rows_units, rows_revenue, rows_cogs = [], [], []
    for row in rows:
        units_arr = np.array(row["monthly_units"], dtype=float)
        demand_mean = float(np.round(units_arr.mean(), 1))
        demand_std = float(np.round(units_arr.std(), 2))
        skus.append(
            {
                "sku_id": row["sku"],
                "name": row["name"],
                "category": row["category"],
                "cost": row["unit_cost"],
                "price": row["unit_price"],
                "unit_margin": round(row["unit_price"] - row["unit_cost"], 4),
                "demand_mean": demand_mean,
                "demand_std": demand_std,
                "elasticity": _elasticity_for(row["category"]),
                "lead_time_days": DEFAULT_LEAD_TIME_DAYS,
                "shelf_space": DEFAULT_SHELF_SPACE,
                "region": IMPORTED_LABEL,
                "channel": IMPORTED_LABEL,
            }
        )
        for mi, month in enumerate(months):
            u = float(units_arr[mi])
            rows_month.append(month)
            rows_sku.append(row["sku"])
            rows_region.append(IMPORTED_LABEL)
            rows_channel.append(IMPORTED_LABEL)
            rows_category.append(row["category"])
            rows_units.append(u)
            rows_revenue.append(round(u * row["unit_price"], 2))
            rows_cogs.append(round(u * row["unit_cost"], 2))

    monthly = {
        "month": np.array(rows_month),
        "sku_id": np.array(rows_sku),
        "region": np.array(rows_region),
        "channel": np.array(rows_channel),
        "category": np.array(rows_category),
        "units": np.array(rows_units, dtype=float),
        "revenue": np.array(rows_revenue, dtype=float),
        "cogs": np.array(rows_cogs, dtype=float),
    }
    ds = Dataset(
        skus=skus,
        monthly=monthly,
        customers=synth.customers,
        route_stops=synth.route_stops,
        depot=synth.depot,
        months=months,
    )
    known = sum(1 for s in skus if s["category"].lower() in _ELASTICITY_BY_CATEGORY)
    assumptions = [
        (
            "customers, RFM segments and delivery routing remain the seeded synthetic "
            "demo layer — the template covers products only"
        ),
        (
            f"price elasticities are category defaults ({known} of {len(skus)} SKUs matched "
            f"a known category; the rest use {DEFAULT_ELASTICITY})"
        ),
        (
            f"lead time is assumed {DEFAULT_LEAD_TIME_DAYS} days and shelf space "
            f"{DEFAULT_SHELF_SPACE} for every imported SKU"
        ),
        (
            "imported data is held in memory only (never written to disk); "
            "restart or POST /api/reset restores the synthetic dataset"
        ),
    ]
    return ds, assumptions
