"""Excel round-trip import + "why this plan?" explanation tests.

Import tests exercise the full loop: template -> fill -> upload -> every
engine (KPIs, plan, exports) runs on the imported data through the same
single-plan path -> the exported workbook traces back to the uploaded values.
Explanation tests pin the maths: contributions sum to the lever totals, the
sensitivity is deterministic, binding constraints are detected correctly on
constructed cases. Every test leaves the app on the synthetic dataset.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app import app as flask_app
from dip import importer


@pytest.fixture
def client():
    flask_app.config["TESTING"] = True
    return flask_app.test_client()


@pytest.fixture(autouse=True)
def _always_restore_synthetic():
    """Whatever a test imports, the next test starts on synthetic data."""
    yield
    flask_app.test_client().post("/api/reset")


def _filled_template(rows: list[tuple]) -> bytes:
    """The real template, filled: (sku, name, category, cost, price, units[24])."""
    wb = load_workbook(BytesIO(importer.build_template()))
    ws = wb["SKUs"]
    for r, (sku, name, category, cost, price, units) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=sku)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=category)
        ws.cell(row=r, column=4, value=cost)
        ws.cell(row=r, column=5, value=price)
        for j, u in enumerate(units):
            ws.cell(row=r, column=6 + j, value=u)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, data: bytes, name: str = "acme.xlsx"):
    return client.post(
        "/api/import",
        data={"workbook": (BytesIO(data), name)},
        content_type="multipart/form-data",
    )


_ROWS = [
    ("ACME-1", "Cola 1L", "Beverages", 1.00, 2.50, [100] * 24),
    ("ACME-2", "Paprika Chips", "Snacks", 0.60, 1.80, [50] * 24),
    ("ACME-3", "Hand Soap", "SomethingNew", 2.00, 5.00, [10] * 24),
]
_ROWS_REVENUE = (100 * 2.50 + 50 * 1.80 + 10 * 5.00) * 24


# ---------------------------------------------------------------------------
# Feature 1: template + import round-trip
# ---------------------------------------------------------------------------


def test_template_ships_instructions_and_exact_header(client):
    resp = client.get("/api/import/template")
    assert resp.status_code == 200
    assert resp.data[:2] == b"PK"
    wb = load_workbook(BytesIO(resp.data))
    assert set(wb.sheetnames) == {"Instructions", "SKUs"}
    header = [c.value for c in wb["SKUs"][1]]
    assert header[:5] == ["sku", "name", "category", "unit_cost", "unit_price"]
    months = header[5:]
    assert len(months) == 24
    assert months == importer.month_columns()
    # the template documents its own rules
    text = "\n".join(str(c.value) for row in wb["Instructions"].iter_rows() for c in row if c.value)
    assert "unit_price" in text and "memory only" in text


def test_import_template_with_data_flows_into_kpis_plan_and_banner(client):
    resp = _upload(client, _filled_template(_ROWS))
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["synthetic"] is False
    assert body["n_skus"] == 3
    assert body["source"]["filename"] == "acme.xlsx"
    assert body["source"]["assumptions"]  # honesty notes ride along

    health = client.get("/api/health").get_json()
    assert health["skus"] == 3
    assert health["source"]["synthetic"] is False

    kpis = client.get("/api/kpis").get_json()
    assert kpis["n_skus"] == 3
    assert abs(kpis["revenue"] - _ROWS_REVENUE) < 0.01

    plan = client.get("/api/prescribe").get_json()
    assert len(plan["cards"]) == 4  # the same plan structure as synthetic mode

    html = client.get("/").data.decode()
    assert "Your data" in html and "acme.xlsx" in html and "3 SKUs" in html
    assert "resetData" in html  # the banner carries the reset button


def test_import_then_export_excel_traces_uploaded_values(client):
    """The round-trip closes: the workbook you download quotes the data you
    uploaded, and its Summary matches the on-screen plan exactly."""
    assert _upload(client, _filled_template(_ROWS)).status_code == 200
    plan = client.get("/api/prescribe").get_json()

    resp = client.get("/api/export/excel")
    wb = load_workbook(BytesIO(resp.data), read_only=True)
    rows = {r[0]: r[1] for r in wb["Summary"].iter_rows(values_only=True) if r and r[0]}
    assert abs(rows["Revenue (24 mo)"] - _ROWS_REVENUE) < 0.01
    assert rows["Expected annual uplift (EUR)"] == plan["expected_uplift_eur"]
    assert "acme.xlsx" in rows["Data source"]
    pricing_skus = {r[0] for r in wb["Pricing"].iter_rows(min_row=2, values_only=True) if r and r[0]}
    assert pricing_skus == {"ACME-1", "ACME-2", "ACME-3"}


def test_imported_mode_scenario_export_matches_screen(client):
    """The existing screen==export consistency guarantee, now in imported mode
    and for a non-default scenario."""
    assert _upload(client, _filled_template(_ROWS)).status_code == 200
    plan = client.get("/api/prescribe?budget=500&max_change=0.05").get_json()
    resp = client.get("/api/export/excel?budget=500&max_change=0.05")
    ws = load_workbook(BytesIO(resp.data), read_only=True)["Summary"]
    rows = {r[0]: r[1] for r in ws.iter_rows(values_only=True) if r and r[0]}
    assert rows["Expected annual uplift (EUR)"] == plan["expected_uplift_eur"]
    assert rows["Uplift % of annual gross margin"] == plan["expected_uplift_pct"]


def test_import_validation_errors_are_cell_addressed_and_atomic(client):
    before = client.get("/api/prescribe").data
    bad_rows = [
        ("DUP-1", "A", "Snacks", 1.0, 2.0, [1] * 24),
        ("DUP-1", "B", "Snacks", 1.0, 2.0, [1] * 24),  # row 3: duplicate sku
        ("BAD-C", "C", "Snacks", "abc", 2.0, [1] * 24),  # row 4: non-numeric cost
        ("NEG-U", "D", "Snacks", 1.0, 2.0, [-5] + [1] * 23),  # row 5: negative units
        ("ZERO-P", "E", "Snacks", 1.0, 0, [1] * 24),  # row 6: price must be > 0
    ]
    resp = _upload(client, _filled_template(bad_rows))
    assert resp.status_code == 400
    errors = resp.get_json()["errors"]
    joined = "\n".join(errors)
    assert "SKUs!A3" in joined and "duplicate" in joined
    assert "SKUs!D4" in joined and "unit_cost" in joined
    assert "SKUs!F5" in joined and ">= 0" in joined
    assert "SKUs!E6" in joined and "> 0" in joined
    # a failed import never touches the running state
    assert client.get("/api/prescribe").data == before
    assert client.get("/api/health").get_json()["skus"] == 200


def test_import_rejects_missing_sheet_wrong_header_and_junk(client):
    # (a) workbook without a SKUs sheet
    wb = Workbook()
    wb.active.title = "Products"
    buf = BytesIO()
    wb.save(buf)
    resp = _upload(client, buf.getvalue())
    assert resp.status_code == 400
    assert "SKUs" in resp.get_json()["error"]

    # (b) tampered header: renamed column + an extra column
    data = _filled_template(_ROWS)
    wb = load_workbook(BytesIO(data))
    ws = wb["SKUs"]
    ws.cell(row=1, column=4, value="cost")  # should be unit_cost
    ws.cell(row=1, column=30, value="notes")  # extra column beyond the template
    buf = BytesIO()
    wb.save(buf)
    resp = _upload(client, buf.getvalue())
    assert resp.status_code == 400
    joined = "\n".join(resp.get_json()["errors"])
    assert "SKUs!D1" in joined and "unit_cost" in joined
    assert "AD1" in joined and "notes" in joined

    # (c) not an xlsx at all
    resp = _upload(client, b"this is not a zip archive", name="junk.xlsx")
    assert resp.status_code == 400
    assert "not a readable" in resp.get_json()["error"]

    # (d) over the documented size cap
    resp = _upload(client, b"0" * (importer.MAX_XLSX_BYTES + 1), name="big.xlsx")
    assert resp.status_code == 400
    assert "MB limit" in resp.get_json()["error"]

    # (e) no file field at all
    resp = client.post("/api/import", data={}, content_type="multipart/form-data")
    assert resp.status_code == 400
    assert "workbook" in resp.get_json()["error"]

    # (f) template with no data rows
    resp = _upload(client, importer.build_template(), name="empty.xlsx")
    assert resp.status_code == 400
    assert "no data rows" in resp.get_json()["error"]

    # none of the failures switched the dataset
    assert client.get("/api/health").get_json()["source"]["synthetic"] is True


def test_reset_restores_byte_identical_synthetic_responses(client):
    kpis_before = client.get("/api/kpis").data
    plan_before = client.get("/api/prescribe").data
    routes_before = client.get("/api/optimize/routes").data

    assert _upload(client, _filled_template(_ROWS)).status_code == 200
    assert client.get("/api/kpis").data != kpis_before  # really switched

    resp = client.post("/api/reset")
    assert resp.status_code == 200
    assert resp.get_json()["synthetic"] is True
    # reset points back at the startup objects: byte-identical responses
    assert client.get("/api/kpis").data == kpis_before
    assert client.get("/api/prescribe").data == plan_before
    assert client.get("/api/optimize/routes").data == routes_before


# ---------------------------------------------------------------------------
# Feature 2: the plan explanation
# ---------------------------------------------------------------------------


def test_explain_contributions_sum_to_lever_totals(client):
    ex = client.get("/api/explain").get_json()
    plan = client.get("/api/prescribe").get_json()
    assert {lv["lever"] for lv in ex["levers"]} == {"pricing", "assortment", "routing"}
    for lv in ex["levers"]:
        recomposed = sum(m["contribution_eur"] for m in lv["moves"]) + lv["remainder_eur"]
        assert abs(recomposed - lv["total_eur"]) < 0.05
        assert lv["total_eur"] == plan["levers"][lv["lever"]]
        assert len(lv["moves"]) <= 5
    assert ex["headline"]["expected_uplift_eur"] == plan["expected_uplift_eur"]


def test_explain_sensitivity_is_deterministic_and_brackets_the_budget(client):
    first = client.get("/api/explain").get_json()
    second = client.get("/api/explain").get_json()
    assert first == second  # two full recomputes, byte-equal numbers
    plan = client.get("/api/prescribe").get_json()
    lo = first["sensitivity"]["budget_minus_10pct"]
    hi = first["sensitivity"]["budget_plus_10pct"]
    assert abs(lo["budget"] - 0.9 * plan["budget"]) < 0.51
    assert abs(hi["budget"] - 1.1 * plan["budget"]) < 0.51
    assert lo["delta_eur"] == round(lo["expected_uplift_eur"] - plan["expected_uplift_eur"], 2)


def test_explain_binding_constraints_on_constructed_cases(client):
    # (a) default synthetic scenario: 40% budget forces SKUs out -> binding;
    #     the +/-15% guardrail clips at least one recommendation -> binding.
    ex = client.get("/api/explain").get_json()
    by_name = {bc["name"]: bc for bc in ex["binding_constraints"]}
    budget_bc = by_name["working-capital budget"]
    guard_bc = next(bc for n, bc in by_name.items() if n.startswith("price guardrail"))
    assert budget_bc["binding"] is True
    assert guard_bc["binding"] is True and guard_bc["hits"] > 0

    # (b) a budget covering the full range cannot bind
    ex = client.get("/api/explain?budget=100000000").get_json()
    budget_bc = next(bc for bc in ex["binding_constraints"] if bc["name"] == "working-capital budget")
    assert budget_bc["binding"] is False

    # (c) constructed slack-guardrail case: one imported Snacks SKU (elasticity
    #     -1.9) priced almost exactly at its Lerner optimum c*e/(e+1) = 2.111 —
    #     no recommendation touches the +/-15% guardrail.
    rows = [("ELAST-1", "Near-optimal", "Snacks", 1.00, 2.11, [10] * 24)]
    assert _upload(client, _filled_template(rows)).status_code == 200
    ex = client.get("/api/explain").get_json()
    guard_bc = next(bc for bc in ex["binding_constraints"] if bc["name"].startswith("price guardrail"))
    assert guard_bc["hits"] == 0
    assert guard_bc["binding"] is False


def test_explain_caveats_follow_the_data_source(client):
    ex = client.get("/api/explain").get_json()
    assert ex["data_source"]["synthetic"] is True
    assert any("synthetic dataset" in c for c in ex["caveats"])

    assert _upload(client, _filled_template(_ROWS)).status_code == 200
    ex = client.get("/api/explain").get_json()
    assert ex["data_source"]["synthetic"] is False
    joined = " ".join(ex["caveats"])
    assert "acme.xlsx" in joined
    assert "category-default elasticities" in joined
    # the routing solved-once note is always present
    assert any("solved once" in c for c in ex["caveats"])


def test_explain_ships_in_excel_and_pdf_exports(client):
    ex = client.get("/api/explain").get_json()

    resp = client.get("/api/export/excel")
    wb = load_workbook(BytesIO(resp.data), read_only=True)
    assert "Explanation" in wb.sheetnames
    ws = wb["Explanation"]
    cells = [c for row in ws.iter_rows(values_only=True) for c in row if c is not None]
    text = "\n".join(str(c) for c in cells)
    assert "Why this plan?" in text
    assert f"EUR {ex['headline']['expected_uplift_eur']:,.0f}/yr" in text
    # the lever totals appear as numbers in the sheet
    for lv in ex["levers"]:
        assert lv["total_eur"] in cells
    # sensitivity budgets too
    assert ex["sensitivity"]["budget_plus_10pct"]["budget"] in cells

    pdf = client.get("/api/export/pdf")
    assert pdf.status_code == 200
    assert pdf.data[:4] == b"%PDF"
    # the deck gained a fifth page (cover, forecast, margin, actions, why)
    assert pdf.data.count(b"/Type /Page ") == 5


def test_dashboard_ships_the_why_panel_and_import_controls(client):
    html = client.get("/").data
    for marker in (b"whyToggle", b"whyBody", b"sec-why", b"importBtn", b"importTemplate", b"importFile"):
        assert marker in html
