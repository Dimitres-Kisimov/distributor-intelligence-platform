"""Scenario A/B compare + Power BI pack tests.

Covers the W3 build: deterministic A/B compare, exact delta arithmetic on a
fixture, the ``POST /api/scenario/compare`` schema + 400 handling, the workbook
Scenarios sheet tracing the compare exactly (screen == export), and the Power BI
star pack (fact sums tie to the KPI headline to the cent; CSVs byte-identical
across runs). Every API test leaves the app on the synthetic dataset.
"""

from __future__ import annotations

import hashlib
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import app as flask_app
from dip import exports, scenario
from dip.data import build_dataset
from dip.optimize import optimize_assortment, optimize_routes
from dip.prescribe import build_plan


@pytest.fixture
def client():
    flask_app.config["TESTING"] = True
    return flask_app.test_client()


@pytest.fixture(autouse=True)
def _always_restore_synthetic():
    """Whatever a test imports, the next test starts on synthetic data."""
    yield
    flask_app.test_client().post("/api/reset")


def _full_capital(ds) -> float:
    return optimize_assortment(ds)["full_capital"]


# ---------------------------------------------------------------------------
# 1. Determinism
# ---------------------------------------------------------------------------


def test_scenario_compare_is_deterministic():
    """Two independent compares of the same scenarios are byte-equal."""
    ds = build_dataset()
    kw = {
        "scenario_a": {"name": "Tight", "budget": 6000.0, "max_change": 0.10},
        "scenario_b": {"name": "Loose", "budget": 20000.0, "max_change": 0.15},
    }
    a = scenario.compare_scenarios(ds, **kw)
    b = scenario.compare_scenarios(ds, **kw)
    assert a == b


# ---------------------------------------------------------------------------
# 2. Delta arithmetic on a fixture (exact numbers)
# ---------------------------------------------------------------------------


def test_delta_math_is_exact_and_ties_to_independent_build_plan():
    ds = build_dataset()
    full = _full_capital(ds)
    budget_a, budget_b = round(0.25 * full, 2), round(0.75 * full, 2)
    res = scenario.compare_scenarios(
        ds,
        scenario_a={"name": "A", "budget": budget_a, "max_change": 0.15},
        scenario_b={"name": "B", "budget": budget_b, "max_change": 0.10},
    )

    # (a) each scenario's KPIs equal an independent, deterministic build_plan
    routes = optimize_routes(ds)
    plan_a = build_plan(ds, budget=budget_a, max_change=0.15, routes=routes)
    plan_b = build_plan(ds, budget=budget_b, max_change=0.10, routes=routes)
    assert res["scenario_a"]["kpis"]["expected_uplift_eur"] == pytest.approx(plan_a["expected_uplift_eur"])
    assert res["scenario_b"]["kpis"]["pricing_uplift_eur"] == pytest.approx(plan_b["levers"]["pricing"])
    assert res["scenario_a"]["kpis"]["assortment_uplift_eur"] == pytest.approx(plan_a["levers"]["assortment"])

    # (b) every delta is EXACTLY B - A (abs) and (B-A)/A (pct), rounded
    for key, d in res["deltas"].items():
        av = res["scenario_a"]["kpis"][key]
        bv = res["scenario_b"]["kpis"][key]
        ndp = 4 if key.endswith("_pct") else 2
        if key == "skus_carried":
            assert d["abs"] == bv - av
        else:
            assert d["abs"] == round(bv - av, ndp)
        assert d["pct"] == (round((bv - av) / av, 4) if av else None)

    # (c) routing is shared, so its lever is identical in both scenarios
    assert res["routing_identical"] is True
    assert res["deltas"]["routing_uplift_eur"]["abs"] == 0.0


def test_more_budget_never_lowers_absolute_margin_captured():
    """Honest sanity check: a bigger budget captures at least as much absolute
    assortment margin (even though the *lever* — the edge over greedy — can
    shrink; that unintuitive effect is documented in the note)."""
    ds = build_dataset()
    full = _full_capital(ds)
    res = scenario.compare_scenarios(
        ds,
        scenario_a={"budget": round(0.25 * full, 2), "max_change": 0.15},
        scenario_b={"budget": round(0.90 * full, 2), "max_change": 0.15},
    )
    assert res["scenario_b"]["kpis"]["margin_captured_eur"] >= res["scenario_a"]["kpis"]["margin_captured_eur"]
    assert "greedy baseline" in res["note"]


# ---------------------------------------------------------------------------
# 3. Parameter validation (unit level)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw",
    [
        {"budget": "abc"},
        {"budget": -5},
        {"budget": 0},
        {"budget": float("inf")},
        {"max_change": 5},
        {"max_change": 0},
        {"max_change": "x"},
        {"name": 123},
        [1, 2, 3],
    ],
)
def test_normalize_scenario_rejects_bad_specs(raw):
    with pytest.raises(scenario.ScenarioError):
        scenario.normalize_scenario(raw, label="A")


def test_normalize_scenario_defaults_and_coercion():
    out = scenario.normalize_scenario(None, label="A")
    assert out == {"name": "A", "budget": None, "max_change": 0.15}
    out = scenario.normalize_scenario({"budget": 6000}, label="B")
    assert out["budget"] == 6000.0 and out["max_change"] == 0.15


# ---------------------------------------------------------------------------
# 4. Endpoint schema + 400
# ---------------------------------------------------------------------------


def test_endpoint_returns_full_schema_and_provenance(client):
    resp = client.post(
        "/api/scenario/compare",
        json={
            "scenario_a": {"name": "Tight", "budget": 6000, "max_change": 0.10},
            "scenario_b": {"name": "Loose", "budget": 20000, "max_change": 0.15},
        },
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert set(body) >= {
        "scenario_a", "scenario_b", "deltas", "provenance", "data_source",
        "routing_identical", "note",
    }
    for side in ("scenario_a", "scenario_b"):
        assert set(body[side]["kpis"]) == set(scenario.KPI_KEYS)
        assert set(body[side]) >= {"name", "budget", "budget_pct_of_full", "max_change", "kpis"}
    assert set(body["deltas"]) == set(scenario.KPI_KEYS)
    # provenance label present and (default state) discloses synthetic data
    assert body["data_source"]["synthetic"] is True
    assert "synthetic" in body["provenance"].lower()


def test_endpoint_deterministic_across_calls(client):
    payload = {"scenario_a": {"budget": 5000, "max_change": 0.15},
               "scenario_b": {"budget": 30000, "max_change": 0.05}}
    a = client.post("/api/scenario/compare", json=payload).data
    b = client.post("/api/scenario/compare", json=payload).data
    assert a == b


def test_endpoint_400_on_bad_params_and_non_json(client):
    # bad budget
    r = client.post("/api/scenario/compare", json={"scenario_a": {"budget": "abc"}, "scenario_b": {}})
    assert r.status_code == 400 and "budget" in r.get_json()["error"]
    # out-of-range guardrail
    r = client.post("/api/scenario/compare", json={"scenario_a": {"max_change": 9}, "scenario_b": {}})
    assert r.status_code == 400 and "max_change" in r.get_json()["error"]
    # non-JSON body
    r = client.post("/api/scenario/compare", data="not json", content_type="text/plain")
    assert r.status_code == 400
    # a bad request never disturbs the served dataset
    assert client.get("/api/health").get_json()["skus"] == 200


# ---------------------------------------------------------------------------
# 5. Scenario sheet: screen == export
# ---------------------------------------------------------------------------


def test_scenario_sheet_traces_the_compare_exactly(client):
    ds = build_dataset()
    full = _full_capital(ds)
    cmp = scenario.compare_scenarios(
        ds,
        scenario_a={"name": "Baseline", "budget": round(0.30 * full, 2), "max_change": 0.15},
        scenario_b={"name": "Stretch", "budget": round(0.85 * full, 2), "max_change": 0.10},
    )
    data = exports.build_excel(ds=ds, scenario_compare=cmp)
    wb = load_workbook(BytesIO(data), read_only=True)
    assert "Scenarios" in wb.sheetnames
    grid = [list(row) for row in wb["Scenarios"].iter_rows(values_only=True)]

    # locate each metric row by its label and assert A / B / Δabs / Δpct cells
    labels = {
        "Expected annual uplift (EUR)": "expected_uplift_eur",
        "Pricing lever (EUR)": "pricing_uplift_eur",
        "Assortment lever vs greedy (EUR)": "assortment_uplift_eur",
        "SKUs carried": "skus_carried",
        "Working-capital budget (EUR)": "budget_eur",
    }
    seen = 0
    for row in grid:
        if row and row[0] in labels:
            key = labels[row[0]]
            assert row[1] == cmp["scenario_a"]["kpis"][key]
            assert row[2] == cmp["scenario_b"]["kpis"][key]
            assert row[3] == cmp["deltas"][key]["abs"]
            assert row[4] == cmp["deltas"][key]["pct"]
            seen += 1
    assert seen == len(labels)


def test_export_endpoint_scenario_sheet_matches_onscreen_and_summary(client):
    """The workbook's Scenarios sheet B-scenario ties to the Summary plan."""
    plan = client.get("/api/prescribe?budget=8000&max_change=0.05").get_json()
    resp = client.get("/api/export/excel?budget=8000&max_change=0.05")
    wb = load_workbook(BytesIO(resp.data), read_only=True)
    summary = {r[0]: r[1] for r in wb["Summary"].iter_rows(values_only=True) if r and r[0]}
    scen = {r[0]: r for r in wb["Scenarios"].iter_rows(values_only=True) if r and r[0]}
    b_uplift = scen["Expected annual uplift (EUR)"][2]  # column C = scenario B
    assert b_uplift == summary["Expected annual uplift (EUR)"] == plan["expected_uplift_eur"]


def test_default_build_excel_ships_scenarios_sheet_with_zero_delta():
    """With no scenario override the default export compares default-vs-default,
    so the sheet exists and every delta is zero."""
    data = exports.build_excel()
    wb = load_workbook(BytesIO(data), read_only=True)
    assert "Scenarios" in wb.sheetnames
    grid = [list(row) for row in wb["Scenarios"].iter_rows(values_only=True)]
    uplift_row = next(r for r in grid if r and r[0] == "Expected annual uplift (EUR)")
    assert uplift_row[3] == 0.0  # Δ abs
    assert uplift_row[4] == 0.0  # Δ %


# ---------------------------------------------------------------------------
# 6. Power BI star pack
# ---------------------------------------------------------------------------


def test_powerbi_fact_sums_tie_to_kpi_headline_to_the_cent():
    from powerbi import build_star

    build_star.build()
    ties = build_star.verify_ties()
    for metric, (fact_sum, headline) in ties.items():
        assert fact_sum == headline, f"{metric}: fact {fact_sum} != headline {headline}"


def test_powerbi_csvs_are_byte_identical_across_runs():
    from powerbi import build_star

    def digests():
        paths = build_star.build()
        return {name: hashlib.sha256(p.read_bytes()).hexdigest() for name, p in paths.items()}

    first = digests()
    second = digests()
    assert first == second
    # the pack ships the required star: fact + dims + kpi_headline
    assert {"fact_sales", "dim_sku", "dim_category", "dim_date", "kpi_headline"} <= set(first)


def test_powerbi_data_is_labelled_synthetic():
    import csv

    from powerbi import build_star

    paths = build_star.build()
    with paths["provenance"].open(encoding="utf-8", newline="") as fh:
        rows = {r["key"]: r["value"] for r in csv.DictReader(fh)}
    assert "SYNTHETIC" in rows["data_label"].upper()


# ---------------------------------------------------------------------------
# 7. Dashboard panel ships
# ---------------------------------------------------------------------------


def test_dashboard_ships_the_scenario_compare_panel(client):
    html = client.get("/").data
    for marker in (b"sec-scenario", b"runScenarioCompare", b"scenarioResult", b"scResDelta"):
        assert marker in html
