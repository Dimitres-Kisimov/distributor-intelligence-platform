"""Flask API tests via ``app.test_client()`` — routes return 200 + JSON keys."""

from __future__ import annotations

import pytest

from app import app as flask_app


@pytest.fixture
def client():
    flask_app.config["TESTING"] = True
    return flask_app.test_client()


def test_health(client):
    resp = client.get("/api/health")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["status"] == "ok"
    assert body["skus"] == 200
    assert body["months"] == 24


def test_kpis_route(client):
    resp = client.get("/api/kpis")
    assert resp.status_code == 200
    body = resp.get_json()
    for key in ("revenue", "gross_margin", "gross_margin_pct", "revenue_series"):
        assert key in body


def test_forecast_route(client):
    resp = client.get("/api/forecast")
    assert resp.status_code == 200
    body = resp.get_json()
    assert "mase" in body
    assert len(body["forecast"]) == 6


def test_prescribe_route(client):
    resp = client.get("/api/prescribe")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["expected_uplift_eur"] > 0
    assert len(body["cards"]) == 4


def test_assortment_route_with_budget(client):
    resp = client.get("/api/optimize/assortment?budget=6000")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["milp"]["margin"] >= body["greedy"]["margin"]


def test_assortment_rejects_non_numeric_budget(client):
    resp = client.get("/api/optimize/assortment?budget=abc")
    assert resp.status_code == 400
    assert "budget" in resp.get_json()["error"]


def test_prices_rejects_non_numeric_max_change(client):
    resp = client.get("/api/optimize/prices?max_change=abc")
    assert resp.status_code == 400
    assert "max_change" in resp.get_json()["error"]


def test_prescribe_respects_budget_and_guardrail(client):
    resp = client.get("/api/prescribe?budget=8000&max_change=0.05")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["budget"] == 8000.0
    assert body["max_change"] == 0.05
    default = client.get("/api/prescribe").get_json()
    # a tighter guardrail must not out-earn the default pricing lever
    assert body["levers"]["pricing"] <= default["levers"]["pricing"]


def test_export_excel_scenario_aware(client):
    resp = client.get("/api/export/excel?budget=8000&max_change=0.05")
    assert resp.status_code == 200
    assert resp.data[:2] == b"PK"


def test_routes_route(client):
    resp = client.get("/api/optimize/routes")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["optimized_km"] <= body["baseline_km"]


def test_routes_prescribe_and_card_quote_same_km(client):
    """One routing solve is the single source of truth: the routes endpoint,
    the prescribe routing lever and the action-card text must all agree."""
    from dip.prescribe import COST_PER_KM, ROUTING_RUNS_PER_YEAR

    routes = client.get("/api/optimize/routes").get_json()
    plan = client.get("/api/prescribe").get_json()
    assert abs(plan["levers"]["routing"] - routes["km_saved"] * COST_PER_KM * ROUTING_RUNS_PER_YEAR) < 0.01
    card = next(c for c in plan["cards"] if c["id"] == "routing")
    assert f"{routes['km_saved']:,.0f} km per run" in card["detail"]


def test_export_excel_matches_onscreen_plan(client):
    """The workbook Summary must quote the exact plan the dashboard serves."""
    from io import BytesIO

    from openpyxl import load_workbook

    plan = client.get("/api/prescribe").get_json()
    resp = client.get("/api/export/excel")
    ws = load_workbook(BytesIO(resp.data), read_only=True)["Summary"]
    rows = {r[0]: r[1] for r in ws.iter_rows(values_only=True) if r and r[0]}
    assert rows["Expected annual uplift (EUR)"] == plan["expected_uplift_eur"]
    assert rows["Uplift % of annual gross margin"] == plan["expected_uplift_pct"]


def test_export_pdf_route(client):
    resp = client.get("/api/export/pdf")
    assert resp.status_code == 200
    assert resp.mimetype == "application/pdf"
    assert resp.data[:4] == b"%PDF"


def test_abc_xyz_route_serves_drill_down_fields(client):
    body = client.get("/api/abc-xyz").get_json()
    row = body["per_sku"][0]
    for key in ("sku_id", "name", "category", "cell", "revenue", "cv"):
        assert key in row


def test_rfm_route_serves_drill_down_fields(client):
    body = client.get("/api/rfm").get_json()
    row = body["per_customer"][0]
    for key in ("customer_id", "name", "segment", "monetary", "recency_months", "frequency"):
        assert key in row


def test_api_open_when_token_unset(client, monkeypatch):
    monkeypatch.delenv("DIP_API_TOKEN", raising=False)
    assert client.get("/api/health").status_code == 200


def test_api_401_when_token_set_and_header_missing_or_wrong(client, monkeypatch):
    monkeypatch.setenv("DIP_API_TOKEN", "s3cret")
    resp = client.get("/api/health")
    assert resp.status_code == 401
    assert "Bearer" in resp.get_json()["error"]
    wrong = client.get("/api/health", headers={"Authorization": "Bearer nope"})
    assert wrong.status_code == 401
    # HTML routes stay open — the guard covers the API surface only
    assert client.get("/").status_code == 200


def test_api_200_with_correct_bearer_token(client, monkeypatch):
    monkeypatch.setenv("DIP_API_TOKEN", "s3cret")
    resp = client.get("/api/health", headers={"Authorization": "Bearer s3cret"})
    assert resp.status_code == 200
    assert resp.get_json()["status"] == "ok"


def test_dashboard_html_renders(client):
    resp = client.get("/")
    assert resp.status_code == 200
    assert b"<html" in resp.data.lower()
    # drill-down panels and the scenario-compare controls ship with the page
    for marker in (b"skuDrill", b"custDrill", b"pinScenario", b"compareBar"):
        assert marker in resp.data
    # the seven designed stations, including the inventory / reliability panels
    # (fed by /api/inventory and /api/reliability), the reconciliation proof
    # strip (fed by /api/reconcile) and the plan-diff bridge (fed by
    # /api/plan-diff) — the change station's own markup is pinned in detail by
    # tests/test_plandiff.py
    for marker in (
        b"st-demand",
        b"st-inventory",
        b"st-reliability",
        b"st-routing",
        b"st-plan",
        b"st-reconcile",
        b"st-change",
        b"invCellTable",
        b"relTable",
        b"proofStrip",
        b"pdBridgeTable",
    ):
        assert marker in resp.data
