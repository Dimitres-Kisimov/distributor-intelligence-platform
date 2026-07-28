"""Cross-sell mining + KPI drill-down tests.

Rule-mining correctness is pinned on a tiny hand-computable basket fixture
(exact support / confidence / lift values, worked out on paper below). The
endpoint tests cover schema, determinism across calls, the honesty labels and
the imported-data (no baskets) path. Drill-down tests prove the segment rows
sum to the exact KPI headline numbers, on screen and in the exported workbook.
"""

from __future__ import annotations

from io import BytesIO

import numpy as np
import pytest
from openpyxl import load_workbook

from app import app as flask_app
from dip import crosssell, importer
from dip.data import SEED, _build_order_lines, build_dataset

# Hand-computable fixture: 8 baskets over items A, B, C.
#   count(A)=5  count(B)=5  count(C)=4
#   count(AB)=3 count(AC)=2 count(BC)=2
# All supports are eighths, so the count/8 divisions are exact in floating
# point and the assertions below can use strict equality.
BASKETS = [
    {"A", "B"},
    {"A", "B"},
    {"A", "B", "C"},
    {"A", "C"},
    {"B"},
    {"C"},
    {"A"},
    {"B", "C"},
]


@pytest.fixture
def client():
    flask_app.config["TESTING"] = True
    return flask_app.test_client()


@pytest.fixture(autouse=True)
def _always_restore_synthetic():
    """Whatever a test imports, the next test starts on synthetic data."""
    yield
    flask_app.test_client().post("/api/reset")


def _filled_template(rows: list[tuple]) -> bytes:
    """The real import template, filled (same shape as test_import_explain)."""
    wb = load_workbook(BytesIO(importer.build_template()))
    ws = wb["SKUs"]
    for r, (sku, name, category, cost, price, units) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=sku)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=category)
        ws.cell(row=r, column=4, value=cost)
        ws.cell(row=r, column=5, value=price)
        for j, u in enumerate(units):
            ws.cell(row=r, column=6 + j, value=u)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, data: bytes, name: str = "acme.xlsx"):
    return client.post(
        "/api/import",
        data={"workbook": (BytesIO(data), name)},
        content_type="multipart/form-data",
    )


# ---------------------------------------------------------------------------
# Rule-mining correctness on the hand-computed fixture
# ---------------------------------------------------------------------------


def test_apriori_exact_supports_on_hand_fixture():
    itemsets = crosssell.apriori(BASKETS, min_support=0.25, max_len=2)
    assert itemsets == {
        frozenset({"A"}): 5 / 8,
        frozenset({"B"}): 5 / 8,
        frozenset({"C"}): 4 / 8,
        frozenset({"A", "B"}): 3 / 8,
        frozenset({"A", "C"}): 2 / 8,
        frozenset({"B", "C"}): 2 / 8,
    }


def test_rules_exact_support_confidence_lift_and_thin_flag():
    itemsets = crosssell.apriori(BASKETS, min_support=0.25, max_len=2)
    rules = crosssell.generate_rules(
        itemsets, len(BASKETS), min_confidence=0.0, min_lift=0.0, thin_support_count=3
    )
    by_pair = {(r["antecedent"], r["consequent"]): r for r in rules}
    assert len(rules) == 6  # both directions of AB, AC, BC

    ab = by_pair[("A", "B")]
    assert ab["support"] == 0.375  # 3 of 8 baskets
    assert ab["confidence"] == 0.6  # 3 of A's 5 baskets
    assert ab["lift"] == 0.96  # 0.6 / (5/8)
    assert ab["support_count"] == 3
    assert ab["thin_support"] is False  # exactly at the threshold

    ca = by_pair[("C", "A")]
    assert ca["confidence"] == 0.5  # 2 of C's 4 baskets
    assert ca["lift"] == 0.8  # 0.5 / (5/8)
    assert ca["support_count"] == 2
    assert ca["thin_support"] is True


def test_rules_filtered_by_confidence_and_lift_thresholds():
    itemsets = crosssell.apriori(BASKETS, min_support=0.25, max_len=2)
    rules = crosssell.generate_rules(
        itemsets, len(BASKETS), min_confidence=0.5, min_lift=0.9
    )
    # Only A<->B survives: C->A and C->B reach confidence 0.5 but lift 0.8.
    assert {(r["antecedent"], r["consequent"]) for r in rules} == {("A", "B"), ("B", "A")}


def test_min_count_for_support_boundaries():
    assert crosssell.min_count_for_support(0.25, 8) == 2
    # float product 0.02 * 6000 lands an ulp above 120 — must not become 121
    assert crosssell.min_count_for_support(0.02, 6000) == 120
    assert crosssell.min_count_for_support(1e-9, 10) == 1  # floor of 1


# ---------------------------------------------------------------------------
# Baskets: reuse of the existing synthetic entities, determinism, no drift
# ---------------------------------------------------------------------------


def test_baskets_reuse_existing_order_events_and_catalogue():
    ds = build_dataset()
    assert ds.order_lines, "synthetic dataset must carry order-line baskets"
    # one basket per order event the RFM history already counts
    assert len(ds.order_lines) == sum(c["frequency"] for c in ds.customers)
    catalogue = {s["sku_id"] for s in ds.skus}
    month_idx = {m: i for i, m in enumerate(ds.months)}
    by_customer = {c["customer_id"]: c for c in ds.customers}
    for order in ds.order_lines:
        assert set(order["sku_ids"]) <= catalogue
        assert len(set(order["sku_ids"])) == len(order["sku_ids"])  # no dup lines
        cust = by_customer[order["customer_id"]]
        assert month_idx[order["month"]] in cust["order_months"]


def test_basket_builder_is_deterministic():
    ds = build_dataset()
    a = _build_order_lines(np.random.default_rng(SEED), ds.skus, ds.customers, ds.months)
    b = _build_order_lines(np.random.default_rng(SEED), ds.skus, ds.customers, ds.months)
    assert a == b


def test_existing_headline_numbers_unchanged_by_basket_generation():
    """The order-line draws are appended to the rng stream, so every number
    published before the cross-sell feature must still be byte-identical."""
    from dip.analytics import kpis

    k = kpis(build_dataset())
    assert k["revenue"] == 4788971.2  # the README's EUR 4,788,971
    assert k["gross_margin"] == 3257507.06
    assert k["gross_margin_pct"] == 0.6802


# ---------------------------------------------------------------------------
# /api/crosssell endpoint
# ---------------------------------------------------------------------------


def test_crosssell_endpoint_schema_and_honesty_label(client):
    resp = client.get("/api/crosssell")
    assert resp.status_code == 200
    body = resp.get_json()
    for key in ("available", "note", "n_baskets", "n_rules", "params", "rules", "products", "source"):
        assert key in body
    assert body["available"] is True
    assert body["source"]["synthetic"] is True
    assert "synthetic" in body["note"] and "not causation" in body["note"]
    rule = body["rules"][0]
    for key in (
        "antecedent", "antecedent_name", "consequent", "consequent_name",
        "support", "confidence", "lift", "support_count", "thin_support",
    ):
        assert key in rule
    assert rule["lift"] >= body["params"]["min_lift"]
    # rules arrive lift-ranked
    lifts = [r["lift"] for r in body["rules"]]
    assert lifts == sorted(lifts, reverse=True)


def test_crosssell_endpoint_deterministic_across_calls(client):
    first = client.get("/api/crosssell?top=50")
    second = client.get("/api/crosssell?top=50")
    assert first.data == second.data  # byte-identical, not just equal-ish


def test_crosssell_product_view_top_n_unknown_404_bad_top_400(client):
    overview = client.get("/api/crosssell").get_json()
    product = overview["products"][0]["sku_id"]
    resp = client.get(f"/api/crosssell?product={product}&top=2")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["product"] == product
    assert 1 <= len(body["recommendations"]) <= 2
    assert all(r["antecedent"] == product for r in body["recommendations"])
    assert "note" in body  # the honesty label rides on every payload

    # a real SKU with no rules is a 200 with an empty list, not an error
    with_rules = {p["sku_id"] for p in overview["products"]}
    quiet = next(s["sku_id"] for s in build_dataset().skus if s["sku_id"] not in with_rules)
    body = client.get(f"/api/crosssell?product={quiet}").get_json()
    assert body["recommendations"] == []

    assert client.get("/api/crosssell?product=SKU-9999").status_code == 404
    resp = client.get("/api/crosssell?top=abc")
    assert resp.status_code == 400
    assert "top" in resp.get_json()["error"]


def test_crosssell_reports_unavailable_on_imported_data(client):
    before = client.get("/api/crosssell?top=50").data
    rows = [("ACME-1", "Cola 1L", "Beverages", 1.00, 2.50, [100] * 24)]
    assert _upload(client, _filled_template(rows)).status_code == 200
    body = client.get("/api/crosssell").get_json()
    assert body["available"] is False
    assert body["n_rules"] == 0 and body["rules"] == []
    assert "basket" in body["note"]  # says WHY: the template has no order lines
    # a product query on imported data stays honest too
    body = client.get("/api/crosssell?product=ACME-1").get_json()
    assert body["available"] is False and body["recommendations"] == []
    # reset restores the startup objects byte-for-byte
    client.post("/api/reset")
    assert client.get("/api/crosssell?top=50").data == before


# ---------------------------------------------------------------------------
# KPI drill-downs: segment rows sum to the exact headline numbers
# ---------------------------------------------------------------------------


def test_kpi_drilldown_sums_tie_to_kpi_headlines(client):
    kpi = client.get("/api/kpis").get_json()
    dd = client.get("/api/kpis/drilldown").get_json()

    seg = dd["revenue_by_segment"]
    assert round(sum(r["revenue"] for r in seg["rows"]), 2) == kpi["revenue"]
    assert seg["total_revenue"] == kpi["revenue"]
    assert len(seg["rows"]) == 15  # 5 regions x 3 channels

    cat = dd["margin_by_category"]
    assert round(sum(r["gross_margin"] for r in cat["rows"]), 2) == kpi["gross_margin"]
    assert cat["total_gross_margin"] == kpi["gross_margin"]
    assert len(cat["rows"]) == 8  # one row per category
    for row in cat["rows"]:
        assert round(row["revenue"] - row["cogs"], 2) == row["gross_margin"]


def test_kpi_drilldown_endpoint_schema_and_determinism(client):
    resp = client.get("/api/kpis/drilldown")
    assert resp.status_code == 200
    body = resp.get_json()
    for key in ("revenue_by_segment", "margin_by_category", "note", "source"):
        assert key in body
    assert "sum to the headline" in body["note"]
    assert client.get("/api/kpis/drilldown").data == resp.data


# ---------------------------------------------------------------------------
# Exports quote the same numbers as the API (screen == export)
# ---------------------------------------------------------------------------


def test_export_excel_crosssell_and_drilldown_sheets_match_api(client):
    cs = client.get("/api/crosssell").get_json()
    kpi = client.get("/api/kpis").get_json()

    wb = load_workbook(BytesIO(client.get("/api/export/excel").data), read_only=True)
    assert "Cross-sell" in wb.sheetnames and "Drill-downs" in wb.sheetnames

    rows = list(wb["Cross-sell"].iter_rows(values_only=True))
    text = " ".join(str(c) for r in rows for c in r if c is not None)
    assert "not causation" in text  # the honesty note ships in the workbook
    header_i = next(i for i, r in enumerate(rows) if r and r[0] == "If they buy")
    top = rows[header_i + 1]
    api_top = cs["rules"][0]
    assert top[0] == api_top["antecedent"] and top[2] == api_top["consequent"]
    assert top[4] == api_top["support"]
    assert top[5] == api_top["confidence"]
    assert top[6] == api_top["lift"]
    assert top[7] == api_top["support_count"]

    rows = list(wb["Drill-downs"].iter_rows(values_only=True))
    totals = [r for r in rows if r and r[0] == "Total"]
    assert totals[0][2] == kpi["revenue"]  # revenue-by-segment total
    assert totals[1][3] == kpi["gross_margin"]  # margin-by-category total


def test_dashboard_ships_crosssell_panel_and_kpi_drilldowns(client):
    html = client.get("/").data
    for marker in (
        b"sec-crosssell", b"crossProduct", b"crossTable",
        b"kpiDrillCard", b"kpiRevenueTile", b"kpiMarginTile",
    ):
        assert marker in html
